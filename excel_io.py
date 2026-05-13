"""
读写 Excel：关键词列 + 回写状态/摘要/截图（嵌入图片）。
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.worksheet import Worksheet


def _script_dir() -> Path:
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def resolve_excel_path(configured: str | None) -> Path:
    if configured and str(configured).strip():
        return Path(configured).expanduser().resolve()
    return _script_dir() / "search_list.xlsx"


def load_rows(
    excel_path: Path,
    sheet_name: str,
    col_keyword: str,
) -> tuple[Worksheet, list[dict[str, Any]], dict[str, int]]:
    """
    返回：worksheet, rows 每项含 _row_index 与 各列单元格对象引用所需数据我们用值读写即可
    这里简化为：list of { '_row': int, 'keyword': str }
    """
    if not excel_path.is_file():
        raise FileNotFoundError(f"未找到 Excel 文件: {excel_path}")

    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表不存在: {sheet_name}，当前有: {wb.sheetnames}")

    ws = wb[sheet_name]
    header_row = 1
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        val = cell.value
        if val is None or str(val).strip() == "":
            continue
        headers[str(val).strip()] = col

    if col_keyword not in headers:
        raise ValueError(f"表头中缺少「{col_keyword}」列，当前表头: {list(headers.keys())}")

    rows: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        kw_cell = ws.cell(row=r, column=headers[col_keyword])
        kw = kw_cell.value
        if kw is None or str(kw).strip() == "":
            continue
        rows.append({"_row": r, "keyword": str(kw).strip()})

    return ws, rows, headers


def write_status(
    ws: Worksheet,
    row: int,
    headers: dict[str, int],
    col_status: str,
    col_result: str,
    status: str,
    result_text: str,
) -> None:
    if col_status in headers:
        ws.cell(row=row, column=headers[col_status], value=status)
    if col_result in headers:
        ws.cell(row=row, column=headers[col_result], value=result_text)


def embed_screenshot(
    ws: Worksheet,
    row: int,
    headers: dict[str, int],
    col_screenshot: str,
    image_path: Path,
    thumb_max_width: int = 480,
) -> None:
    if col_screenshot not in headers:
        return
    col_idx = headers[col_screenshot]
    cell = ws.cell(row=row, column=col_idx)

    # 清理旧图片：openpyxl 对同单元格多图支持弱，这里只写文字占位+新图
    img = XLImage(str(image_path))
    if img.width and img.width > thumb_max_width:
        ratio = thumb_max_width / float(img.width)
        img.width = int(img.width * ratio)
        img.height = int(img.height * ratio)

    anchor = f"{cell.column_letter}{row}"
    img.anchor = anchor
    ws.add_image(img)


def save_workbook(ws: Worksheet, excel_path: Path) -> None:
    wb = ws.parent
    # 写到临时文件再替换，降低写入中断导致损坏的概率
    tmp = excel_path.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    tmp.replace(excel_path)


def ensure_output_dir() -> Path:
    d = _script_dir() / "output" / "screenshots"
    d.mkdir(parents=True, exist_ok=True)
    return d
